class Process:
    __numberOfProcesses = 1

    def __init__(self, startTime, time, num):
        self.__id = num
        if time < 1:
            time = 1
        self.__startTime = startTime  # czas przybycia
        self.__timeToTheEnd = time  # czas trwania
        self.__burstTime = time
        self.__waitingTime = 0  # czas oczekiwania
        self.__executeTime = 0  # czas przybycia
        self.__isFinished = False  # czy zakonczony
        self.__isActive = False

    def __str__(self):
        template = "ID: {} \nStart time={} [ms] \nBurst time={} [ms]\nWaiting time={} [ms]\nExecute Time={} [ms]\nis finished={} \nis active={}\n"
        return template.format(self.__id, self.__startTime, self.__burstTime, self.__waitingTime, self.__executeTime,
                               self.__isFinished, self.__isActive)

    # metoda zwiększająca czas oczekiwania
    def increaseWaitingTime(self):
        if not self.__isFinished and not self.__isActive:
            self.__waitingTime += 1

    # metoda obliczająca całkowity czas trwania
    def setExecuteTime(self, presentTime):
        if self.__isFinished:
            self.__executeTime = presentTime - self.__startTime

    # metoda zmniejszająca czas wykonywania
    def decreaseBurstTime(self):
        self.__timeToTheEnd -= 1
        if self.__timeToTheEnd == 0:
            self.__isFinished = True
            self.__isActive = False


class RoundRobinProcess(Process):
    def __init__(self, startTime, time, num):
        super().__init__(startTime, time, num)

    def setExecuteTime(self, presentTime):
        self._Process__executeTime = presentTime - self._Process__startTime - self._Process__timeToTheEnd


# sort by start
# sortowanie wg rozpoczęcia
def sortProcessessStartTime(processes):
    for i in range(len(processes)):
        for k in range(len(processes)):
            if processes[i]._Process__startTime < processes[k]._Process__startTime:
                processes[k], processes[i] = processes[i], processes[k]


# checking if all processes have been handled
# sprawdzanie czy wszystkie procesy zostały obsłużone
def isFinished(active, all):
    if len(active) == 0 and all[0]._Process__isFinished == True:
        for process in all:
            if process._Process__isFinished == False:
                return False
        return True
    return False


# function listing processes
# funckja wyświetlająca listę procesów
def printArrayOfProcesses(array):
    for process in array:
        print(process)


def printGraph(graph):
    for i in graph:
        print(i, end=' ')


# a function that counts the average execution time
# funckja licząca średni czas wykonania
def executeTimeAverage(processes):
    if len(processes) == 0:
        return
    executeTime = 0
    for process in processes:
        executeTime += process._Process__executeTime
    return executeTime / len(processes)


# a function that counts the average execution time
# funckja licząca średni czas wykonania
def waitingTimeAverage(processes):
    if len(processes) == 0:
        return
    waitingTime = 0
    for process in processes:
        waitingTime += process._Process__waitingTime
    return waitingTime / len(processes)


def summary(processes):
    print('Summary: ')
    # printArrayOfProcesses(processes)
    print('Average waiting time: ', waitingTimeAverage(processes))
    print('Average Executive Time: ', executeTimeAverage(processes))

# Shortest Remaining Time First (SRTF)
def srtf(processes):
    sortProcessessStartTime(processes)
    time = processes[0]._Process__startTime
    for process in processes:
        process._Process__startTime -= time
    time = 0
    activeProcesses = []
    graph = []
    while True:
        # adding new processes to the process table on the processor
        # dodawanie nowych procesów do tablicy procesów na procesorze
        for process in processes:
            if time == process._Process__startTime:
                activeProcesses.append(process)
        # sorting an array of processes on the processor
        # posortowanie tablicy procesów na procesorze
        activeProcesses.sort(key=lambda x: x._Process__timeToTheEnd)
        time += 1
        if len(activeProcesses) > 0:
            # reduction of execution time on the current process
            # increase the total execution time
            # zmniejszenie czasu wykoniania na obecnym procesie
            # zwiększenie czasu całkowitego wykonania
            if activeProcesses[0]._Process__isActive == False:
                _Process__isActive = True
                activeProcesses[0].decreaseBurstTime()
                # graph += ('|    ' + str(activeProcesses[0]._Process__id) + '    |')
                graph.append(str(activeProcesses[0]._Process__id))

            # zwiększenie czasu całkowitego wykonania
            # zwiększenie czasu czekania
            for i in range(1, len(activeProcesses)):
                activeProcesses[i]._Process__isActive = False
                activeProcesses[i].increaseWaitingTime()
            # sprawdzenie czy proces wykonywany się nie zakończył
            if activeProcesses[0]._Process__isFinished:
                activeProcesses[0].setExecuteTime(time)
                del activeProcesses[0]

        if isFinished(activeProcesses, processes):
            break
        if len(activeProcesses) == 0:
            print('--waiting--')
    return processes, graph


# Round Robin
def roundRobin(processes, quantum):
    sortProcessessStartTime(processes)
    if len(processes) == 0:
        return
    time = processes[0]._Process__startTime
    for process in processes:
        process._Process__startTime -= time
    time = 0
    activeProcesses = []
    quantumCopy = quantum
    graph = []
    while True:
        # dodawanie nowych procesów do tablicy procesów na procesorze
        for process in processes:
            if time == process._Process__startTime:
                activeProcesses.append(process)
        quantumCopy -= 1
        time += 1
        if len(activeProcesses) > 0:
            # ustawienie odpowiednich flag
            if not activeProcesses[0]._Process__isActive:
                activeProcesses[0]._Process__isActive = True
            for idx in range(1, len(activeProcesses)):
                activeProcesses[idx]._Process__isActive = False

            # zmniejszenie czasu wykonywania
            activeProcesses[0].decreaseBurstTime()
            graph.append(str(activeProcesses[0]._Process__id))
            # zwiększenie czasu oczekiwania
            for process in activeProcesses:
                if process._Process__isActive == False:
                    process.increaseWaitingTime()

            if activeProcesses[0]._Process__isFinished:
                quantumCopy = quantum
                activeProcesses[0].setExecuteTime(time)
                del activeProcesses[0]

        if quantumCopy == 0:
            if len(activeProcesses) > 0:
                activeProcesses.append(activeProcesses[0])
                del activeProcesses[0]
            quantumCopy = quantum

        if isFinished(activeProcesses, processes):
            break
    return processes, graph

# Shortest Job First
def sjf(processes):
    sortProcessessStartTime(processes)
    time = processes[0]._Process__startTime
    for process in processes:
        process._Process__startTime -= time
    time = 0
    activeProcesses = []
    graph = []
    while True:
        # dodawanie nowych procesów do tablicy procesów na procesorze
        for process in processes:
            if time == process._Process__startTime:
                activeProcesses.append(process)
        if time == 0:
            activeProcesses.sort(key=lambda x: x._Process__timeToTheEnd)
        time += 1
        if len(activeProcesses) > 0:
            # zmniejszenie czasu wykoniania na obecnym procesie
            # zwiększenie czasu całkowitego wykonania
            if activeProcesses[0]._Process__isActive == False:
                _Process__isActive = True
                activeProcesses[0].decreaseBurstTime()
                graph.append(str(activeProcesses[0]._Process__id))

            # zwiększenie czasu całkowitego wykonania
            # zwiększenie czasu czekania
            for i in range(1, len(activeProcesses)):
                activeProcesses[i]._Process__isActive = False
                activeProcesses[i].increaseWaitingTime()
            # sprawdzenie czy proces wykonywany się nie zakończył
            if activeProcesses[0]._Process__isFinished:
                activeProcesses[0].setExecuteTime(time)
                del activeProcesses[0]
                # posortowanie tablicy procesów na procesorze
                activeProcesses.sort(key=lambda x: x._Process__timeToTheEnd)

        if isFinished(activeProcesses, processes):
            break
        if len(activeProcesses) == 0:
            print('--waiting--')
    return processes, graph


# First Came First Serve
def fcfs(processes):
    sortProcessessStartTime(processes)
    if len(processes) == 0:
        return
    time = processes[0]._Process__startTime
    for process in processes:
        process._Process__startTime -= time
    time = 0
    activeProcesses = []
    graph = []
    while True:
        for process in processes:
            if time == process._Process__startTime:
                activeProcesses.append(process)
        time += 1
        if len(activeProcesses) > 0:
            if not activeProcesses[0]._Process__isActive:
                activeProcesses[0]._Process__isActive = True
            activeProcesses[0].decreaseBurstTime()
            graph.append(str(activeProcesses[0]._Process__id))
            for i in range(1, len(activeProcesses)):
                activeProcesses[i]._Process__isActive = False
                activeProcesses[i].increaseWaitingTime()
            # sprawdzenie czy proces wykonywany się nie zakończył
            if activeProcesses[0]._Process__isFinished:
                activeProcesses[0].setExecuteTime(time)
                del activeProcesses[0]

        if isFinished(activeProcesses, processes):
            break
        if len(activeProcesses) == 0:
            print('--waiting--')
    return processes, graph


import openpyxl as opxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import COLOR_INDEX
import os, json
from string import ascii_uppercase
from random import randint


def importData(processesTable, names,graphes, fileName, isPainting):
    file = opxl.Workbook()
    p=0
    for package in processesTable:
        processes=package
        graph=graphes[p]
        sheet = file.create_sheet(names[p])
        p += 1
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Burst time [ms]'
        sheet['C1'] = 'Start time [ms]'
        sheet['D1'] = 'Waiting time [ms]'
        sheet['E1'] = 'Executive time [ms]'
        i = 2
        for process in processes:
            sheet['A' + str(i)] = process._Process__id
            sheet['B' + str(i)] = process._Process__burstTime
            sheet['C' + str(i)] = process._Process__startTime
            sheet['D' + str(i)] = process._Process__waitingTime
            sheet['E' + str(i)] = process._Process__executeTime
            i += 1
        sheet['G1'] = 'Average waiting time [ms]'
        sheet['G2'] = 'Average executive time wania [ms]'
        sheet['H1'] = waitingTimeAverage(processes)
        sheet['H2'] = executeTimeAverage(processes)
        if isPainting:
            j =0
            m = 8
            colours = list(COLOR_INDEX)
            del colours[0]
            fits = {}
            for k in range(len(graph)):
                if j == 0:
                    sheet[str(ascii_uppercase[m]) + str(i + 5)] = graph[k]
                    if graph[k] not in fits.keys():
                        idx = randint(0, len(list(COLOR_INDEX)) - 2)
                        sheet[str(ascii_uppercase[m]) + str(i + 5)].fill = PatternFill(start_color=colours[idx],
                                                                                       fill_type='solid')
                        fits[graph[k]] = colours[idx]
                    else:
                        idx = (fits[graph[k]])
                        sheet[str(ascii_uppercase[m]) + str(i + 5)].fill = PatternFill(start_color=idx,
                                                                                       fill_type='solid')
                else:
                    sheet[str(ascii_uppercase[j - 1]) + str(ascii_uppercase[m]) + str(i + 5)] = graph[k]
                    if graph[k] not in fits.keys():
                        idx = randint(0, len(list(COLOR_INDEX)) - 2)
                        sheet[str(ascii_uppercase[j - 1]) + str(ascii_uppercase[m]) + str(i + 5)].fill = PatternFill(
                            start_color=colours[idx], fill_type='solid')
                        fits[graph[k]] = colours[idx]
                    else:
                        idx = (fits[graph[k]])
                        sheet[str(ascii_uppercase[j - 1]) + str(ascii_uppercase[m]) + str(i + 5)].fill = PatternFill(
                            start_color=idx, fill_type='solid')

                m += 1
                if m != 0 and m % len(ascii_uppercase) == 0:
                    j += 1
                    m = 0

    file.save(fileName)


def saveToFile(processes, graph, name, fileName, isPainting=False):
    while True:
        try:
            importData(processes, graph, str(name), fileName, isPainting)
            break
        except PermissionError:
            fileName = input('Niedozwolona operacja- brak dostępu do tego pliku. Podaj nową nazwę pliku: ')
            if not fileName[-5:] == '.xlsx':
                fileName += '.xlsx'


from os import strerror
def returnTimesFromFile(filePath, sheetName, column):
    try:
        sheet =opxl.load_workbook(filePath)[sheetName]
        # sheet=file[sheetName]
        myList=[]
        for i in range(1,sheet.max_row+1):
            row = [cell.value for cell in sheet[i]][column]
            try:
                myList.append(int(row))
            except:
                continue
        return myList
    except IOError as e:
        print('Błąd I/O:',strerror(e.errno))


def exportList(pages, fileName):
    with open(fileName, 'w') as filehandle:
        json.dump(pages, filehandle)


def importList(fileName):
    with open(fileName, 'r') as filehandle:
        pages = json.load(filehandle)
    return pages